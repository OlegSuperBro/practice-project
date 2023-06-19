import psycopg2
import string
import openpyxl
import calendar
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from psycopg2.extras import DictCursor, DictRow
from typing import List, Union

from datatypes import User, Vacation, Post
from config import CONFIG


class Verifier:
    @staticmethod
    def password_is_valid(password: str) -> bool:
        if len(password) < CONFIG.minimal_password_length:
            return False

        if password in CONFIG.banned_passwords:
            return False

        # check if password using only english letters and ascii symbols
        if not all([x in string.printable[:-6] for x in password]):
            return False

        return True

    @staticmethod
    def username_is_valid(username: str) -> bool:
        if len(username) < CONFIG.minimal_username_length:
            return False
        if username in CONFIG.banned_usernames:
            return False

        # check if username using only english letters and ascii symbols
        if not all([x in string.printable[:-6] for x in username]):
            return False

        return True

    def vacation_is_valid(vacation: Vacation) -> bool:
        if vacation.start_date >= vacation.end_date:
            return False

        return True


def get_connection():
    """
    Returns connection to database

    Returns
    ----
    psycopg2.connection - connection to database
    None - No connection :(  (error accured, possibly database is offline)
    """
    try:
        return psycopg2.connect(dbname="vacations", user="postgres", password="root")
    except psycopg2.DatabaseError:
        return None


def ping_connection():
    if get_connection() is None:
        return False
    return True


def export_to_excel() -> openpyxl.Workbook:
    all_user = DataBase().get_users()

    workbook = openpyxl.Workbook()
    worksheet: Worksheet = workbook.active

    worksheet.merge_cells(None, 1, 2, 1, len(calendar.month_name) - 1)
    worksheet.cell(1, 2, "Месяцы")
    worksheet["B1"].alignment = Alignment(horizontal="center")
    worksheet.append(list(calendar.month_name))

    worksheet.merge_cells(None, 1, 1, 2, 1)
    worksheet.cell(1, 1, "ФИО")
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.append(list(calendar.month_name))

    column_widths = []

    for user in all_user:
        tmp_username = f"{user.post.name.capitalize() if user.post is not None else ''}\n{user.first_name if user.first_name is not None else ''} {user.second_name if user.second_name is not None else ''} {user.surname if user.second_name is not None else ''}"
        tmp_row = [tmp_username, [], [], [], [], [], [], [], [], [], [], [], []]
        for vacation in user.vacations:
            tmp_row[vacation.start_date.month].append(f"{vacation.start_date.strftime('%d/%m/%Y')} - {vacation.end_date.strftime('%d/%m/%Y')}")

        for index, tmp_list in enumerate(tmp_row):
            if isinstance(tmp_list, list):
                tmp_row[index] = "\n".join(tmp_list)[1:]
        worksheet.append(tmp_row)

        for i, cell in enumerate(tmp_row):
            if len(column_widths) > i:
                if len(cell.split("\n")[0]) > column_widths[i]:
                    column_widths[i] = len(cell.split("\n")[0])
            else:
                column_widths += [len(cell.split("\n")[0])]

    for i, column_width in enumerate(column_widths, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = column_width
    worksheet.column_dimensions[get_column_letter(1)].width = 30

    worksheet.delete_rows(3)

    return workbook


class DataBase:
    def __init__(self, connection=get_connection()) -> None:
        self.connection = connection

    def user_exist(self, user: Union[User, str]):
        if isinstance(user, User):
            user = user.login
        with self.connection.cursor(cursor_factory=DictCursor) as cur:
            cur.execute(f"SELECT * FROM users WHERE login='{user}'")
            return len(cur.fetchall()) != 0

    def register_user(self, user: User) -> None:
        with self.connection.cursor(cursor_factory=DictCursor) as cur:
            cur.execute("INSERT INTO users (login, password, first_name, second_name, surname, post, role) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                        (user.login, user.password, user.first_name, user.second_name, user.surname, user.post.name, user.role))
            self.connection.commit()

        return 0

    def login_user(self, user: User) -> bool:
        with self.connection.cursor(cursor_factory=DictCursor) as cur:
            cur.execute(f"SELECT * FROM users WHERE login='{user.login}' AND password='{user.password}'")
            return len(cur.fetchall()) != 0

    def check_recovery_for_user(self, user: User, code: int) -> bool:
        with self.connection.cursor(cursor_factory=DictCursor) as cur:
            cur.execute(f"SELECT * FROM users WHERE recovery_code={code} AND login='{user.login}'")
            return len(cur.fetchall()) > 0

    def get_vacations_for_user(self, user: User) -> List[Vacation]:
        with self.connection.cursor(cursor_factory=DictCursor) as cur:
            cur.execute(f"SELECT * FROM vacations WHERE user_id={user._id}")
            result = cur.fetchall()
            return [Vacation.from_sql(**vacation) for vacation in result]

    def get_post(self, name: str) -> Union[Post, None]:
        """
        Get post from it name

        Args
        ----
        name: str
            name of post

        Returns
        ----
        Post
            post representation
        """
        with self.connection.cursor(cursor_factory=DictCursor) as cur:
            cur.execute(f"SELECT * FROM posts WHERE name='{name}'")
            result = cur.fetchone()
            if result is None:
                return None
            return Post.from_sql(**result)

    def get_posts(self) -> Union[list[Post], None]:
        with self.connection.cursor(cursor_factory=DictCursor) as cur:
            cur.execute("SELECT * FROM posts")
            result = cur.fetchall()
            return [Post.from_sql(**data) for data in result]

    def get_recovery_for_user(self, user: User) -> int:
        with self.connection.cursor(cursor_factory=DictCursor) as cur:
            cur.execute(f"SELECT recovery_code FROM users WHERE login='{user.login}'")
            return cur.fetchone()[0]

    def get_user(self, user: Union[str, User]) -> Union[User, None]:
        if isinstance(user, User):
            user = user.login
        with self.connection.cursor(cursor_factory=DictCursor) as cur:
            cur.execute(f"SELECT * FROM users WHERE login='{user}'")
            result = cur.fetchone()
            result["post"] = self.get_post(result.get("post"))
            return User.from_sql(**result, vacations=self.get_vacations_for_user(User(_id=result.get("id"))))

    def get_users(self) -> Union[List[User], None]:
        with self.connection.cursor(cursor_factory=DictCursor) as cur:
            cur.execute("SELECT * FROM users")
            result = cur.fetchall()
            for index, value in enumerate(result):
                result[index]["post"] = self.get_post(value.get("post"))
            return [User.from_sql(**data, vacations=self.get_vacations_for_user(User(_id=data.get("id")))) for data in result]

    def add_vacation(self, user: User, vacation: Vacation, commit: bool = True) -> None:
        with self.connection.cursor(cursor_factory=DictCursor) as cur:
            cur.execute("INSERT INTO vacations (user_id, start_date, end_date) VALUES (%s, %s, %s)",
                        (user._id, vacation.start_date.strftime("%d/%m/%Y"), vacation.end_date.strftime("%d/%m/%Y")))
            if commit:
                self.commit()

    def delete_vacations_for_user(self, user: User, commit: bool = True) -> None:
        with self.connection.cursor(cursor_factory=DictCursor) as cur:
            cur.execute(f"DELETE FROM vacations WHERE user_id='{user._id}'")
            if commit:
                self.commit()

    def delete_recovery_for_user(self, user: User, commit: bool = True) -> None:
        with self.connection.cursor(cursor_factory=DictCursor) as cur:
            cur.execute(f"UPDATE users SET recovery_code=null WHERE login='{user.login}'")
            if commit:
                self.commit()

    def delete_user(self, user: User, commit: bool = True) -> None:
        self.delete_vacations_for_user(user, commit)
        with self.connection.cursor(cursor_factory=DictCursor) as cur:
            cur.execute(f"DELETE FROM users WHERE login='{user.login}'")
            if commit:
                self.commit()

    def set_recovery_for_user(self, user: User, code: int, commit: bool = True) -> None:
        with self.connection.cursor(cursor_factory=DictCursor) as cur:
            cur.execute(f"UPDATE users SET recovery_code={code} WHERE login='{user.login}'")
            if commit:
                self.commit()

    def set_password_for_user(self, user: User, password: int, commit: bool = True) -> None:
        with self.connection.cursor(cursor_factory=DictCursor) as cur:
            cur.execute(f"UPDATE users SET password='{password}' WHERE login='{user.login}'")
            if commit:
                self.commit()

    def commit(self) -> None:
        self.connection.commit()

    def execute_sql(self, command: str) -> List[DictRow]:
        with self.connection.cursor(cursor_factory=DictCursor) as cur:
            cur.execute(command)
            return cur.fetchall()

    def __delete__(self, instance):
        self.connection.close()


if __name__ == "__main__":
    # user = DataBase().get_user("root")
    # print(user)
    # vacations = DataBase().get_vacations_for_user(user)
    # print(vacations)

    export_to_excel()
